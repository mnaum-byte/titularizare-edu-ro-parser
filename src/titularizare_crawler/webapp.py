from flask import Flask
from flask import Response, request
import threading
import uuid
import time


def create_app() -> Flask:
    app = Flask(__name__)
    # In-memory job store (simple, non-persistent)
    jobs: dict[str, dict] = {}

    def new_job_id() -> str:
        return uuid.uuid4().hex

    @app.get("/")
    def index():
        return (
            """
            <!doctype html>
            <html lang="en">
            <head>
              <meta charset="utf-8" />
              <meta name="viewport" content="width=device-width, initial-scale=1" />
              <title>titularizare.edu.ro crawler</title>
              <link rel="preconnect" href="https://fonts.googleapis.com">
              <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
              <link href="https://fonts.googleapis.com/css2?family=Roboto:wght@400;500;700&display=swap" rel="stylesheet">
              <link href="https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined:opsz,wght,FILL,GRAD@24,400,0,0" rel="stylesheet" />
              <script type="module">import 'https://esm.sh/@material/web@1.4.1/all.js';</script>
              <style>
                :root { color-scheme: dark; --md-sys-color-surface:#121212; --md-sys-color-on-surface:#E4E4E7; --md-sys-color-surface-variant:#1E1E1E; --md-sys-color-surface-container:#18181B; --md-sys-color-outline:#3F3F46; }
                body { font-family: Roboto, system-ui, -apple-system, Segoe UI, Arial, sans-serif; margin:0; background:var(--md-sys-color-surface); color:var(--md-sys-color-on-surface); }
                .container { width:max-content; max-width:100%; margin:12vh auto; padding:0 16px; }
                .title { font-size:28px; font-weight:700; margin-bottom:24px; letter-spacing:.2px; }
                .card { border-radius:16px; padding:20px; border:1px solid var(--md-sys-color-outline); background:var(--md-sys-color-surface-variant); display:flex; align-items:center; flex-wrap:nowrap; gap:12px; }
                md-menu-anchor { display:inline-block; position:relative; }
                .combo { position:relative; flex:1 1 580px; min-width:460px; }
                .chips-infield { position:absolute; top:14px; left:12px; right:64px; display:flex; gap:6px; flex-wrap:wrap; pointer-events:none; }
                .chips-infield > * { pointer-events:auto; }
                .clear-chip { position:absolute; top:50%; transform:translateY(-50%); right:8px; pointer-events:auto; z-index:3; }
                .clear-chip::part(label) { display:inline-block; margin:0; font-size:20px; line-height:1; }
                .clear-chip::part(container) { padding-inline:0; gap:0; }
                .match { background: rgba(138,180,248,.2); padding:0 2px; border-radius:3px; }
                md-menu:not(:defined), md-menu-item:not(:defined) { display:none; }
                md-menu {
                  --md-menu-container-color: var(--md-sys-color-surface-container);
                  --md-menu-item-label-text-color: var(--md-sys-color-on-surface);
                }
                md-menu-item {
                  --md-menu-item-label-text-color: var(--md-sys-color-on-surface);
                }
                /* Dialog theming and sizing (scoped to app dialogs only) */
                .app-dialog {
                  --md-dialog-container-color: var(--md-sys-color-surface-variant);
                  color: var(--md-sys-color-on-surface);
                  min-width: 720px !important;
                  min-height: 640px !important;
                  width: auto !important;
                  height: auto !important;
                  min-inline-size: 720px !important;
                  min-block-size: 640px !important;
                }
                /* Some Material versions size the "surface" part; cover both */
                .app-dialog::part(container),
                .app-dialog::part(surface) {
                  min-width: 720px !important;
                  min-height: 640px !important;
                  min-inline-size: 720px !important;
                  min-block-size: 640px !important;
                  max-width: calc(100vw - 32px);
                  max-height: calc(100vh - 32px);
                  max-inline-size: calc(100vi - 32px);
                  max-block-size: calc(100vb - 32px);
                }
                .app-dialog::part(content) { min-block-size: 640px; }
                .dlg-head { display:flex; align-items:center; justify-content:space-between; gap:16px; }
                .dlg-summary { margin: 0; font-size: 14px; color: #c9c9ce; }
                /* Draw a divider between headline and content to mirror actions divider */
                .app-dialog [slot="content"] {
                  border-top: 1px solid rgba(255,255,255,.16);
                  margin-top: 8px;
                  padding-top: 12px;
                }
                .app-dialog .dlg-actions { display:flex; align-items:center; gap:12px; }
                .app-dialog .dlg-actions md-linear-progress { flex:1; display:none; }
                /* Dialog lists - dark styling */
                .app-dialog md-list {
                  --md-list-container-color: var(--md-sys-color-surface-container);
                  background: var(--md-sys-color-surface-container);
                  border: 1px solid var(--md-sys-color-outline);
                  border-radius: 12px;
                  padding: 4px;
                  margin: 8px 0 16px;
                }
                .app-dialog md-list-item {
                  --md-list-item-label-text-color: var(--md-sys-color-on-surface);
                  --md-list-item-container-color: transparent;
                }
                .app-dialog md-list-item::part(container) { border-radius: 8px; }
                .app-dialog md-list-item:hover::part(container) { background: rgba(255,255,255,.06); }
                .app-dialog md-list-item:focus-visible::part(container) { outline: 2px solid #8AB4F8; outline-offset: 2px; }
                .app-dialog md-list-item:not(.list-title)::part(container) { padding-inline-start: 0; }
                .app-dialog md-list-item:not(.list-title) div[slot="headline"] { padding-inline-start: 14px; }
                .app-dialog md-list-item .export-btn { margin-inline-start: 12px; }
                .app-dialog .export-btn {
                  --md-outlined-button-outline-color: rgba(255,255,255,.28);
                  --md-outlined-button-label-text-color: #fff;
                  --md-outlined-button-disabled-outline-color: rgba(255,255,255,.12);
                  --md-outlined-button-disabled-label-text-color: rgba(255,255,255,.38);
                }
                .app-dialog .export-list-btn {
                  --md-filled-button-container-color: rgba(138,180,248,.18);
                  --md-filled-button-label-text-color: #E4E4E7;
                  margin-inline-start: 12px;
                }
                .app-dialog md-list-item.list-title::part(container) {
                  background: transparent;
                  border-radius: 8px 8px 0 0;
                  padding-block: 6px;
                  margin-bottom: 0;
                }
                .app-dialog md-list-item.list-title div[slot="headline"] {
                  font-weight: 700;
                  color: var(--md-sys-color-on-surface);
                  opacity: .9;
                }
                /* Two-list layout: contiguous outer border with inner divider */
                .app-dialog md-list.list-title-only {
                  margin-bottom: 0;
                  border-bottom-left-radius: 0;
                  border-bottom-right-radius: 0;
                  border-bottom: 0; /* let items list draw the inner divider */
                }
                .app-dialog md-list.list-items {
                  margin-top: 0;
                  border-top-left-radius: 0;
                  border-top-right-radius: 0;
                  /* draw the inner divider using the top border of the items list */
                  border-top-color: rgba(255,255,255,.16);
                  border-top-width: 1px;
                  border-top-style: solid;
                }
                /* Remove label transition to avoid flicker when floating */
                md-outlined-text-field::part(label) { transition: none !important; }
                /* Snackbar: bottom-centered and above other UI */
                #appSnackbar {
                  position: fixed;
                  inset-inline: 0;
                  inset-block-end: calc(24px + env(safe-area-inset-bottom, 0px));
                  display: flex;
                  justify-content: center;
                  z-index: 10000;
                  /* Ensure visible pill styling via component tokens */
                  --md-snackbar-container-color: rgba(40,40,44,.96);
                  --md-snackbar-label-text-color: #FFFFFF;
                  /* Disable native snackbar display now that we use stacked toasts */
                  display: none !important;
                }
                /* Cover different implementations */
                #appSnackbar::part(container) {
                  border-radius: 8px;
                }
                #appSnackbar::part(surface) {
                  background: var(--md-sys-color-surface-container);
                  color: var(--md-sys-color-on-surface);
                  border-radius: 8px;
                  box-shadow: 0 4px 12px rgba(0,0,0,.5);
                  padding-inline: 16px;
                }
                /* Fallback toasts (stackable) */
                #appToasts {
                  position: fixed;
                  inset-inline: 0;
                  inset-block-end: calc(24px + env(safe-area-inset-bottom, 0px));
                  display: flex;
                  flex-direction: column;
                  align-items: center;
                  justify-content: flex-end; /* newest at the bottom */
                  gap: 8px;
                  z-index: 10001;
                  pointer-events: none;
                }
                .appToast {
                  background: var(--md-sys-color-surface-container);
                  color: var(--md-sys-color-on-surface);
                  border-radius: 8px;
                  padding: 10px 16px;
                  box-shadow: 0 4px 12px rgba(0,0,0,.5);
                  opacity: 0;
                  transform: translateY(8px);
                  transition: opacity .18s ease, transform .18s ease;
                  font: 500 14px/20px Roboto, system-ui, -apple-system, Segoe UI, Arial, sans-serif;
                  pointer-events: none;
                  max-width: min(92vw, 720px);
                }
                .appToast.show { opacity: 1; transform: translateY(0); }
              </style>
            </head>
            <body>
              <main class="container">
                <div class="title">titularizare.edu.ro crawler</div>
                <div class="card">
                  <md-menu-anchor id="judeteAnchor">
                    <div class="combo">
                      <md-outlined-text-field id="judeteField" label="Județe" role="combobox" aria-autocomplete="list" aria-haspopup="menu" aria-expanded="false" aria-controls="judeteMenu" style="width:100%">
                        <div id="summaryText" slot="supporting-text"></div>
                        <md-icon id="expandIcon" slot="trailing-icon">expand_more</md-icon>
                      </md-outlined-text-field>
                      <div id="chipsInField" class="chips-infield"></div>
                      <md-assist-chip id="clearAllChip" class="clear-chip" label="×" title="Golește selecția" aria-label="Golește selecția"></md-assist-chip>
                    </div>
                    <md-menu id="judeteMenu"></md-menu>
                  </md-menu-anchor>
                  <md-filled-button id="extractBtn" style="margin-left:12px;" disabled>Extrage Candidati</md-filled-button>
                  <md-filled-button id="extractPostsBtn" style="margin-left:12px;" disabled>Extrage Posturi</md-filled-button>
                </div>
              </main>
              <md-dialog id="dialogCandidati" class="app-dialog">
                <div slot="headline" class="dlg-head"><span>Extrage Candidati</span><md-icon-button id="dlgCandClose" aria-label="Închide"><md-icon>close</md-icon></md-icon-button></div>
                <div slot="content">
                  <p id="dlgCandSummary" class="dlg-summary"></p>
                  <div id="dlgCandLists"></div>
                </div>
                <div slot="actions" class="dlg-actions">
                  <md-linear-progress id="dlgCandProgress" four-color></md-linear-progress>
                  <md-filled-button id="dlgCandExportAll" class="export-list-btn">Export all</md-filled-button>
                </div>
              </md-dialog>
              <md-dialog id="dialogPosturi" class="app-dialog">
                <div slot="headline" class="dlg-head"><span>Extrage Posturi</span><md-icon-button id="dlgPostClose" aria-label="Închide"><md-icon>close</md-icon></md-icon-button></div>
                <div slot="content">
                  <p id="dlgPostSummary" class="dlg-summary"></p>
                  <div id="dlgPostLists"></div>
                </div>
                <div slot="actions" class="dlg-actions">
                  <md-linear-progress id="dlgPostProgress" four-color></md-linear-progress>
                  <md-filled-button id="dlgPostExportAll" class="export-list-btn">Export all</md-filled-button>
                </div>
              </md-dialog>
              <md-snackbar id="appSnackbar">
                <div id="snackbarLabel" slot="label"></div>
              </md-snackbar>
              <div id="appToasts" role="status" aria-live="polite"></div>
              <script>
                const field = document.getElementById('judeteField');
                const menu = document.getElementById('judeteMenu');
                const anchorEl = document.getElementById('judeteAnchor');
                const chips = document.getElementById('chipsInField');
                const extractBtn = document.getElementById('extractBtn');
                const extractPostsBtn = document.getElementById('extractPostsBtn');
                const dialogCandidati = document.getElementById('dialogCandidati');
                const dialogPosturi = document.getElementById('dialogPosturi');
                const dlgCandSummary = document.getElementById('dlgCandSummary');
                const dlgPostSummary = document.getElementById('dlgPostSummary');
                const dlgPostLists = document.getElementById('dlgPostLists');
                const dlgCandLists = document.getElementById('dlgCandLists');
                const dlgCandExportAll = document.getElementById('dlgCandExportAll');
                const dlgPostExportAll = document.getElementById('dlgPostExportAll');
                const dlgCandProgress = document.getElementById('dlgCandProgress');
                const dlgPostProgress = document.getElementById('dlgPostProgress');
                const appSnackbar = document.getElementById('appSnackbar');
                const snackbarLabel = document.getElementById('snackbarLabel');
                const appToasts = document.getElementById('appToasts');
                let loaded = false; let options = []; const selected = new Set();
                let menuScrollTop = 0; let suppressOpenUntil = 0;
                let exportingAllCand = false; let exportingAllPost = false;
                const Job = {
                  async start(payload){
                    const resp = await fetch('/api/export/jobs', {method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify(payload)});
                    if(!resp.ok) throw new Error('job start failed');
                    return await resp.json(); // {id}
                  },
                  async status(id){ const r = await fetch(`/api/export/jobs/${id}`); return await r.json(); },
                  async download(id){ return `/api/export/jobs/${id}/download`; }
                };
                async function runJobAndDownload(payload, progressEl){
                  const TAIL_MS = 8000; let rafId=0;
                  if(progressEl){ progressEl.style.display='block'; progressEl.indeterminate=true; progressEl.value=0; }
                  const {id} = await Job.start(payload);
                  let done=false, hadTotal=false;
                  while(!done){
                    const s = await Job.status(id);
                    if(progressEl && s.total>0){ progressEl.indeterminate=false; hadTotal=true; progressEl.value = Math.min(1, s.done / s.total); }
                    if(s.status==='ready'){ done=true; break; }
                    if(s.status==='error'){ throw new Error(s.error||'error'); }
                    await new Promise(r=>setTimeout(r, 500));
                  }
                  const url = await Job.download(id);
                  const a=document.createElement('a'); a.href=url; a.download=(payload.filename||'export.xlsx'); document.body.appendChild(a); a.click(); a.remove();
                  if(progressEl){ const startVal = progressEl.indeterminate?0:(progressEl.value||0); const startTs = performance.now(); const step=()=>{ const p=Math.min(1,(performance.now()-startTs)/TAIL_MS); progressEl.indeterminate=false; progressEl.value=Math.min(1,startVal+(1-startVal)*p); if(p<1) rafId=requestAnimationFrame(step); else progressEl.style.display='none'; }; rafId=requestAnimationFrame(step); }
                }

                function syncLabelFloat(){
                  const hasSel = selected.size > 0;
                  const isFocused = document.activeElement === field;
                  if (isFocused) {
                    if (field.value === ' ') field.value = '';
                    return;
                  }
                  if (hasSel) {
                    if (field.value === '') field.value = ' ';
                  } else {
                    if (field.value === ' ') field.value = '';
                  }
                }
                function getLabelForCode(code){ const it = menu.querySelector(`[data-code="${code}"]`); return (it && it.getAttribute('data-label')) || code; }
                function showSnackbar(message){
                  try {
                    if (snackbarLabel) snackbarLabel.textContent = message;
                    // Native snackbar disabled; rely on stacked toasts
                    if (appToasts) {
                      const t = document.createElement('div');
                      t.className = 'appToast';
                      t.textContent = message;
                      appToasts.appendChild(t);
                      // animate in next frame
                      requestAnimationFrame(()=> t.classList.add('show'));
                      // auto-dismiss after 5s
                      setTimeout(()=>{
                        t.classList.remove('show');
                        setTimeout(()=>{ t.remove(); }, 220);
                      }, 5000);
                    }
                  } catch (e) { console.warn('snackbar', e); }
                }

                async function downloadWithProgress(url, filename, progressEl){
                  const TAIL_MS = 8000; let finishAt=null; let rafId=0;
                  try{
                    if(progressEl){ progressEl.style.display='block'; progressEl.indeterminate = true; progressEl.value = 0; }
                    const resp = await fetch(url);
                    if(!resp.ok){ throw new Error(`HTTP ${resp.status}`); }
                    const total = Number(resp.headers.get('Content-Length')||0);
                    const reader = resp.body && resp.body.getReader ? resp.body.getReader() : null;
                    if(reader && total>0 && progressEl){ progressEl.indeterminate=false; progressEl.value = 0; }
                    let received = 0; const chunks=[];
                    if(reader){
                      while(true){
                        const {done, value} = await reader.read();
                        if(done) break;
                        chunks.push(value);
                        received += value.length;
                        if(progressEl && total>0){ progressEl.value = Math.min(1, received/total); }
                      }
                      const blob = new Blob(chunks, {type: 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'});
                      const urlB = URL.createObjectURL(blob);
                      const a=document.createElement('a'); a.href=urlB; a.download=filename; document.body.appendChild(a); a.click(); a.remove(); URL.revokeObjectURL(urlB);
                      if(progressEl){ const startVal = progressEl.indeterminate?0:(progressEl.value||0); const startTs = performance.now(); const step=()=>{ const p=Math.min(1,(performance.now()-startTs)/TAIL_MS); progressEl.indeterminate=false; progressEl.value=Math.min(1,startVal+(1-startVal)*p); if(p<1) rafId=requestAnimationFrame(step); }; rafId=requestAnimationFrame(step); finishAt = startTs + TAIL_MS; }
                    } else {
                      const blob = await resp.blob();
                      const urlB = URL.createObjectURL(blob);
                      const a=document.createElement('a'); a.href=urlB; a.download=filename; document.body.appendChild(a); a.click(); a.remove(); URL.revokeObjectURL(urlB);
                      if(progressEl){ const startVal = progressEl.indeterminate?0:(progressEl.value||0); const startTs = performance.now(); const step=()=>{ const p=Math.min(1,(performance.now()-startTs)/TAIL_MS); progressEl.indeterminate=false; progressEl.value=Math.min(1,startVal+(1-startVal)*p); if(p<1) rafId=requestAnimationFrame(step); }; rafId=requestAnimationFrame(step); finishAt = startTs + TAIL_MS; }
                    }
                  } finally {
                    if(progressEl){ const hide=()=>{ if(rafId) cancelAnimationFrame(rafId); progressEl.style.display='none'; }; const delay = finishAt? Math.max(0, finishAt - performance.now()) : 8000; setTimeout(hide, delay); }
                  }
                }
                function updateFieldLabel(){ const arr=[...selected]; const labels = arr.map(c=>getLabelForCode(c)); const summary = labels.length? labels.slice(0,3).join(', ') + (labels.length>3? ` +${labels.length-3}`:'') : 'Neselectat'; const st=document.getElementById('summaryText'); if(st) st.textContent = summary; syncLabelFloat(); const disabled = selected.size === 0; if (extractBtn) extractBtn.disabled = disabled; if (extractPostsBtn) extractPostsBtn.disabled = disabled; }
                function renderChips(){
                  chips.innerHTML='';
                  const arr=[...selected];
                  arr.slice(0,3).forEach(code=>{
                    const chip=document.createElement('md-input-chip');
                    chip.setAttribute('label', getLabelForCode(code));
                    chip.removable=true;
                    chip.addEventListener('remove',()=>{ selected.delete(code); const it=menu.querySelector(`[data-code="${code}"]`); if(it) it.selected=false; renderChips(); updateFieldLabel(); });
                    chip.tabIndex = 0;
                    chips.appendChild(chip);
                  });
                  // Attach Export All (Candidati) once
                  if (dlgCandExportAll && !dlgCandExportAll.__boundOnce) {
                    dlgCandExportAll.__boundOnce = true;
                    dlgCandExportAll.addEventListener('click', async ()=>{
                      const codes=[...selected]; if(codes.length===0) return; if(exportingAllCand) return; exportingAllCand=true;
                      const labels=codes.map(c=>getLabelForCode(c)); const safe=(s)=> (s||'').replace(/[^\w\s\-_,.()]/g,'').trim().slice(0,120)||'export';
                      const filename = `${safe(`Export Candidati - ${labels.join(', ')}`)}.xlsx`; showSnackbar(`Export mixt pornit: ${labels.join(', ')}`);
                      try{ await runJobAndDownload({type:'candidati', codes, labels, filename}, dlgCandProgress);}catch(e){ console.error(e);} finally {exportingAllCand=false;}
                    });
                  }
                  // Attach Export All (Posturi) once (ensure bound regardless of filter usage)
                  if (dlgPostExportAll && !dlgPostExportAll.__boundOnce) {
                    dlgPostExportAll.__boundOnce = true;
                    dlgPostExportAll.addEventListener('click', async ()=>{
                      const codes=[...selected]; if(codes.length===0) return; if(exportingAllPost) return; exportingAllPost=true;
                      const labels=codes.map(c=>getLabelForCode(c)); const safe=(s)=> (s||'').replace(/[^\w\s\-_,.()]/g,'').trim().slice(0,120)||'export';
                      const filename = `${safe(`Export Posturi - ${labels.join(', ')}`)}.xlsx`; showSnackbar(`Export mixt pornit: ${labels.join(', ')}`);
                      try{ await runJobAndDownload({type:'posturi', codes, labels, filename}, dlgPostProgress);}catch(e){ console.error(e);} finally {exportingAllPost=false;}
                    });
                  }
                  if(arr.length>3){ const more=document.createElement('md-assist-chip'); more.setAttribute('label', `+${arr.length-3}`); chips.appendChild(more); }
                  updateFieldLabel();
                  // Toggle clear chip, expand icon, and button state after any selection change
                  const clearChip = document.getElementById('clearAllChip');
                  const expandIcon = document.getElementById('expandIcon');
                  const hasSel = selected.size>0;
                  if (clearChip) clearChip.style.display = hasSel ? '' : 'none';
                  if (expandIcon) {
                    if (hasSel) { expandIcon.style.display='none'; expandIcon.setAttribute('hidden',''); }
                    else { expandIcon.style.display=''; expandIcon.removeAttribute('hidden'); }
                  }
                  const chipsBox = document.getElementById('chipsInField');
                  if (chipsBox) chipsBox.style.right = hasSel ? '12px' : '64px';
                  if (extractBtn) extractBtn.disabled = !hasSel;
                  if (extractPostsBtn) extractPostsBtn.disabled = !hasSel;
                }
                function populateMenu(values){ menu.innerHTML='';
                  const actionAll=document.createElement('md-menu-item');
                  // Keep menu open when activating this action
                  actionAll.keepOpen = true; actionAll.setAttribute('keepOpen','');
                  let l=document.createElement('div'); l.slot='headline'; l.textContent='Selectează toate'; actionAll.appendChild(l);
                  actionAll.addEventListener('click', (e)=>{ e.preventDefault(); e.stopPropagation(); const keepTop = menuScrollTop; const visible = Array.from(menu.querySelectorAll('md-menu-item')).filter(it=>it.hasAttribute('data-code') && it.style.display!=='none'); const toSelect = visible.length? visible : Array.from(menu.querySelectorAll('md-menu-item')).filter(it=>it.hasAttribute('data-code')); toSelect.forEach(it=>{ it.selected=true; const code=it.getAttribute('data-code'); if(code) selected.add(code); }); renderChips(); menu.open=true; requestAnimationFrame(()=>{ menu.scrollTop = keepTop; }); });
                  menu.appendChild(actionAll);
                  values.forEach(opt=>{ const code = typeof opt === 'string' ? opt : (opt.code||''); const text = typeof opt === 'string' ? opt : (opt.label||opt.code||''); const item=document.createElement('md-menu-item');
                    item.type='checkbox';
                    // Keep menu open when toggling items
                    item.keepOpen = true; item.setAttribute('keepOpen','');
                    item.setAttribute('data-code', code);
                    item.setAttribute('data-label', text);
                    const label=document.createElement('div'); label.slot='headline'; label.textContent=text; label.setAttribute('data-text', text); item.appendChild(label);
                    item.addEventListener('click',(ev)=>{ ev.preventDefault(); ev.stopPropagation(); const keepTop = menuScrollTop; item.selected=!item.selected; if(item.selected){ if(code) selected.add(code);} else { if(code) selected.delete(code);} renderChips(); requestAnimationFrame(()=>{ menu.scrollTop = keepTop; }); }); menu.appendChild(item); }); }
                function filterMenu(q){
                  q=(q||'').trim().toUpperCase();
                  menu.querySelectorAll('md-menu-item').forEach(it=>{
                    const code=(it.getAttribute('data-code')||'').toUpperCase();
                    const text=(it.getAttribute('data-label')||'').toUpperCase();
                    const show=q? (code.includes(q) || text.includes(q)) : true;
                    it.style.display=show? '':'none';
                    const label=it.querySelector('[slot="headline"]');
                    if(label){
                      const original=label.getAttribute('data-text')||label.textContent||'';
                      if(!q){ label.textContent=original; }
                      else {
                        const hay = original.toUpperCase();
                        const idx=hay.indexOf(q);
                        label.innerHTML= idx>=0? `${original.slice(0,idx)}<span class=\"match\">${original.slice(idx,idx+q.length)}</span>${original.slice(idx+q.length)}` : original;
                      }
                    }
                  });
                  // Attach Export All (Posturi) once
                  if (dlgPostExportAll && !dlgPostExportAll.__boundOnce) {
                    dlgPostExportAll.__boundOnce = true;
                    dlgPostExportAll.addEventListener('click', async ()=>{
                      const codes=[...selected]; if(codes.length===0) return; if(exportingAllPost) return; exportingAllPost=true;
                      const labels=codes.map(c=>getLabelForCode(c)); const safe=(s)=> (s||'').replace(/[^\w\s\-_,.()]/g,'').trim().slice(0,120)||'export';
                      const filename = `${safe(`Export Posturi - ${labels.join(', ')}`)}.xlsx`; showSnackbar(`Export mixt pornit: ${labels.join(', ')}`);
                      try{ await runJobAndDownload({type:'posturi', codes, labels, filename}, dlgPostProgress);}catch(e){ console.error(e);} finally {exportingAllPost=false;}
                    });
                  }
                  const active = document.activeElement && document.activeElement.closest ? document.activeElement.closest('md-menu-item') : null;
                  if(!active || active.style.display==='none'){
                    const first = Array.from(menu.querySelectorAll('md-menu-item')).find(it=>it.style.display!== 'none' && it.hasAttribute('data-code'));
                    if(first) first.focus();
                  }
                }
                async function ensureLoaded(){ if(loaded) return; loaded=true; try{ const res=await fetch('/api/judete'); const data=await res.json(); options=(data.items||data.values||[]); populateMenu(options);} catch(e){ loaded=false; console.error(e);} }
                function forceLabelFloatForOpen(){ if(field.value==='') field.value=' '; }
                function restoreLabelAfterClose(){ if(selected.size===0) { if(field.value===' ') field.value=''; } else { if(field.value==='') field.value=' '; } }
                function isSuppressed(){ return performance.now() < suppressOpenUntil; }
                function suppressOpen(ms=300){ suppressOpenUntil = performance.now() + ms; }
                async function openMenuSafely(){
                  if(isSuppressed()) return;
                  if(!anchorEl || !anchorEl.isConnected) return;
                  await ensureLoaded();
                  await customElements.whenDefined('md-menu');
                  if(!menu || !menu.isConnected) return;
                  try{
                    menu.anchorElement = anchorEl;
                    forceLabelFloatForOpen();
                    menu.open = true;
                    field.setAttribute('aria-expanded','true');
                  }catch(e){ console.warn('menu/open', e); }
                }
                function focusFirstVisibleMenuItem(){ const first = Array.from(menu.querySelectorAll('md-menu-item')).find(it=>it.style.display!== 'none' && it.hasAttribute('data-code')); if(first) first.focus(); }
                function getVisibleMenuItems(){ return Array.from(menu.querySelectorAll('md-menu-item')).filter(it=>it.style.display!== 'none' && it.hasAttribute('data-code')); }
                function focusRelativeMenuItem(delta){
                  const items=getVisibleMenuItems();
                  if(items.length===0) return false;
                  const active=document.activeElement && document.activeElement.closest ? document.activeElement.closest('md-menu-item') : null;
                  let idx = items.indexOf(active);
                  if(idx === -1){
                    // If starting from a non-data item (e.g., action), go to first/last accordingly
                    const startIdx = delta >= 0 ? 0 : items.length - 1;
                    items[startIdx].focus();
                    return true;
                  }
                  const next = items[idx + delta];
                  if(next){ next.focus(); return true; }
                  return false;
                }
                function getFirstChip(){ return chips.querySelector('md-input-chip'); }
                function hasChips(){ return !!getFirstChip(); }
                function focusFirstChip(){ const c=getFirstChip(); if(c) c.focus(); }
                (async function initUI(){ await ensureLoaded(); await customElements.whenDefined('md-menu'); menu.anchorElement = anchorEl; try{ menu.closeOnSelect=false; }catch(e){}; menu.style.maxHeight='480px'; menu.style.minHeight='360px'; menu.style.overflowY='auto';
                  // Keep label floated while menu is open; restore after close based on selection
                  menu.addEventListener('opened', ()=>{ if(field.value==='') field.value=' '; field.setAttribute('aria-expanded','true'); });
                  menu.addEventListener('closed', ()=>{ if(selected.size===0 && field.value===' ') field.value=''; field.setAttribute('aria-expanded','false'); });
                  // Pre-float label on pointerdown to avoid any visual jump
                  anchorEl.addEventListener('pointerdown', ()=>{ if(field.value==='') field.value=' '; }, {capture:true});
                  field.addEventListener('focus', ()=>{ openMenuSafely(); });
                  field.addEventListener('click', ()=>{ openMenuSafely(); });
                  document.querySelector('.combo')?.addEventListener('click', ()=>{ openMenuSafely(); });
                  const clearChip = document.getElementById('clearAllChip');
                  const expandIcon = document.getElementById('expandIcon');
                  function syncClearChip(){
                    const hasSel = selected.size>0;
                    // Show clear chip only when there are selections
                    clearChip.style.display = hasSel ? '' : 'none';
                    // Fully remove expand icon from layout when clear chip is visible
                    if (expandIcon) {
                      if (hasSel) { expandIcon.style.display = 'none'; expandIcon.setAttribute('hidden',''); }
                      else { expandIcon.style.display = ''; expandIcon.removeAttribute('hidden'); }
                    }
                    // Allow chips to use the freed trailing space when expand icon is hidden
                    const chipsBox = document.getElementById('chipsInField');
                    if (chipsBox) chipsBox.style.right = hasSel ? '12px' : '64px';
                    if (extractBtn) extractBtn.disabled = !hasSel;
                    if (extractPostsBtn) extractPostsBtn.disabled = !hasSel;
                  }
                  clearChip.addEventListener('keydown', async (e)=>{
                    if(e.key==='Tab' && !e.shiftKey){ e.preventDefault(); await openMenuSafely(); requestAnimationFrame(()=>{ focusFirstVisibleMenuItem(); }); }
                  });
                  clearChip.addEventListener('click', (e)=>{ e.stopPropagation(); suppressOpen(350); const keepTop = menuScrollTop; selected.clear(); menu.querySelectorAll('md-menu-item').forEach(it=>{ if(it.hasAttribute('data-code')) it.selected=false; }); renderChips(); updateFieldLabel(); syncClearChip(); if(menu.open){ menu.open=false; field.setAttribute('aria-expanded','false'); restoreLabelAfterClose(); } requestAnimationFrame(()=>{ menu.scrollTop = keepTop; }); });
                  syncClearChip();
                  // Open simple Material dialogs with only a title
                  extractBtn?.addEventListener('click', async ()=>{
                    await customElements.whenDefined('md-dialog');
                    const codes=[...selected]; const labels = codes.map(c=>getLabelForCode(c));
                    if (dlgCandSummary) dlgCandSummary.textContent = labels.length ? `Județe selectate: ${labels.join(', ')}` : 'Nu ați selectat niciun județ.';
                    if (dlgCandLists) {
                      dlgCandLists.innerHTML = '';
                      const codeToLabel = new Map(codes.map(c=>[c, getLabelForCode(c)]));
                      for (const code of codes){
                        const label = codeToLabel.get(code) || code;
                        const container = document.createElement('div');
                        const listTitle = document.createElement('md-list');
                        listTitle.classList.add('list-title-only');
                        // Title row as a non-interactive list item
                        const titleItem = document.createElement('md-list-item');
                        titleItem.classList.add('list-title');
                        titleItem.tabIndex = -1;
                        // Keep title row non-focusable, but allow child button clicks
                        titleItem.style.pointerEvents = 'auto';
                        const th = document.createElement('div'); th.slot = 'headline'; th.textContent = label; titleItem.appendChild(th);
                        const exportAll = document.createElement('md-filled-button');
                        exportAll.className = 'export-list-btn';
                        exportAll.textContent = `Export ${label}`;
                        exportAll.slot = 'end';
                        exportAll.addEventListener('click', async (ev)=>{ ev.stopPropagation(); try{ const clean=(s)=> (s||'').replace(/[^\w\s\-_,.()]/g,' ').replace(/\s{2,}/g,' ').trim().slice(0,120)||'Export Candidati'; const filename = `${clean('Export Candidati ' + label)}.xlsx`; showSnackbar(`Export listă pornit: ${label}`); await downloadWithProgress(`/api/export/judet-bulk?code=${encodeURIComponent(code)}&type=candidati&sheetBase=${encodeURIComponent(label)}&filename=${encodeURIComponent(filename)}`, filename, dlgCandProgress);}catch(e){ console.error(e);} });
                        titleItem.appendChild(exportAll);
                        listTitle.appendChild(titleItem);
                        const list = document.createElement('md-list');
                        list.classList.add('list-items');
                        // Loading row placeholder
                        // Initially empty; we'll populate after fetch
                        container.appendChild(listTitle);
                        container.appendChild(list);
                        dlgCandLists.appendChild(container);
                        try {
                          const res = await fetch(`/api/judete/${code}/candidati-options`);
                          const data = await res.json();
                          // Clear current list content and populate with fetched items
                          list.innerHTML = '';
                          (data.items||[]).forEach((entry)=>{
                            const it = document.createElement('md-list-item');
                            const d = document.createElement('div'); d.slot='headline'; d.textContent = entry.label || entry.value || '';
                            it.appendChild(d);
                            const btn = document.createElement('md-outlined-button');
                            btn.className = 'export-btn';
                            btn.textContent = 'Export';
                            btn.slot = 'end';
                            btn.addEventListener('click', async (ev)=>{ ev.stopPropagation(); try{ const sheet = `${label} - ${(entry.value||'').trim()}`; const safe = (s)=> (s||'').replace(/[^a-z0-9_\-]+/gi,'').slice(0,40) || 'item'; const filename = `export_${safe(code)}_${safe(entry.value)}.xlsx`; showSnackbar(`Export pornit: ${sheet}`); await downloadWithProgress(`/api/export/judet?code=${encodeURIComponent(code)}&value=${encodeURIComponent(entry.value||'')}&sheet=${encodeURIComponent(sheet)}&filename=${encodeURIComponent(filename)}`, filename, dlgCandProgress);}catch(e){ console.error(e);} });
                            it.appendChild(btn);
                            list.appendChild(it);
                          });
                          if(list.children.length===0){ const empty=document.createElement('md-list-item'); const dd=document.createElement('div'); dd.slot='headline'; dd.textContent='Niciun element.'; empty.appendChild(dd); list.appendChild(empty); }
                        } catch (e) {
                          list.innerHTML = '';
                          const err = document.createElement('md-list-item'); const de=document.createElement('div'); de.slot='headline'; de.textContent='Eroare la încărcare.'; err.appendChild(de); list.appendChild(err);
                        }
                      }
                    }
                    if (dialogCandidati) dialogCandidati.open = true;
                  });
                  extractPostsBtn?.addEventListener('click', async ()=>{
                    await customElements.whenDefined('md-dialog');
                    const codes=[...selected]; const labels = codes.map(c=>getLabelForCode(c));
                    if (dlgPostSummary) dlgPostSummary.textContent = labels.length ? `Județe selectate: ${labels.join(', ')}` : 'Nu ați selectat niciun județ.';
                    if (dlgPostLists) {
                      dlgPostLists.innerHTML = '';
                      const codeToLabel = new Map(codes.map(c=>[c, getLabelForCode(c)]));
                      for (const code of codes){
                        const label = codeToLabel.get(code) || code;
                        const container = document.createElement('div');
                        const listTitle = document.createElement('md-list');
                        listTitle.classList.add('list-title-only');
                        const titleItem = document.createElement('md-list-item');
                        titleItem.classList.add('list-title');
                        titleItem.tabIndex = -1; titleItem.style.pointerEvents = 'auto';
                        const th = document.createElement('div'); th.slot = 'headline'; th.textContent = label; titleItem.appendChild(th);
                        const exportAll = document.createElement('md-filled-button');
                        exportAll.className = 'export-list-btn'; exportAll.textContent = `Export ${label}`; exportAll.slot='end';
                        exportAll.addEventListener('click', async (ev)=>{ ev.stopPropagation(); try{ const clean=(s)=> (s||'').replace(/[^\w\s\-_,.()]/g,' ').replace(/\s{2,}/g,' ').trim().slice(0,120)||'Export Posturi'; const filename = `${clean('Export Posturi ' + label)}.xlsx`; showSnackbar(`Export listă pornit: ${label}`); await downloadWithProgress(`/api/export/judet-bulk?code=${encodeURIComponent(code)}&type=posturi&sheetBase=${encodeURIComponent(label)}&filename=${encodeURIComponent(filename)}`, filename, dlgPostProgress);}catch(e){ console.error(e);} });
                        titleItem.appendChild(exportAll);
                        listTitle.appendChild(titleItem);
                        const list = document.createElement('md-list'); list.classList.add('list-items');
                        container.appendChild(listTitle); container.appendChild(list);
                        dlgPostLists.appendChild(container);
                        try {
                          const res = await fetch(`/api/judete/${code}/posturi-options`);
                          const data = await res.json();
                          list.innerHTML = '';
                          (data.items||[]).forEach((entry)=>{
                            const it = document.createElement('md-list-item');
                            const d = document.createElement('div'); d.slot='headline'; d.textContent = entry.label || entry.value || '';
                            it.appendChild(d);
                            const btn = document.createElement('md-outlined-button');
                            btn.className = 'export-btn'; btn.textContent = 'Export'; btn.slot = 'end';
                            btn.addEventListener('click', async (ev)=>{ ev.stopPropagation(); try{ const sheet = `${label} - ${(entry.value||'').trim()}`; const safe = (s)=> (s||'').replace(/[^a-z0-9_\-]+/gi,'').slice(0,40) || 'item'; const filename = `export_posturi_${safe(code)}_${safe(entry.value)}.xlsx`; showSnackbar(`Export pornit: ${sheet}`); await downloadWithProgress(`/api/export/judet?code=${encodeURIComponent(code)}&value=${encodeURIComponent(entry.value||'')}&sheet=${encodeURIComponent(sheet)}&filename=${encodeURIComponent(filename)}`, filename, dlgPostProgress);}catch(e){ console.error(e);} });
                            it.appendChild(btn);
                            list.appendChild(it);
                          });
                          if(list.children.length===0){ const empty=document.createElement('md-list-item'); const dd=document.createElement('div'); dd.slot='headline'; dd.textContent='Niciun element.'; empty.appendChild(dd); list.appendChild(empty); }
                        } catch (e) {
                          list.innerHTML = '';
                          const err = document.createElement('md-list-item'); const de=document.createElement('div'); de.slot='headline'; de.textContent='Eroare la încărcare.'; err.appendChild(de); list.appendChild(err);
                        }
                      }
                    }
                    if (dialogPosturi) dialogPosturi.open = true;
                  });
                  document.getElementById('dlgCandClose')?.addEventListener('click', ()=>{ if(dialogCandidati) dialogCandidati.open = false; });
                  document.getElementById('dlgPostClose')?.addEventListener('click', ()=>{ if(dialogPosturi) dialogPosturi.open = false; });
                  // Track scroll to preserve position through UI updates
                  menu.addEventListener('scroll', ()=>{ menuScrollTop = menu.scrollTop; });
                  // Close on outside click using composed path to avoid false positives
                  document.addEventListener('click', (e)=>{
                    if(!menu.open) return;
                    const path = e.composedPath ? e.composedPath() : [];
                    const inMenu = path.includes(menu) || menu.contains(e.target);
                    const inAnchor = path.includes(anchorEl) || anchorEl.contains(e.target);
                    if(!inMenu && !inAnchor){
                      menu.open=false; field.setAttribute('aria-expanded','false'); restoreLabelAfterClose();
                    }
                  });
                  // Tab from field: move to first chip if any, otherwise into the menu
                  field.addEventListener('keydown', async (e)=>{ if(e.key==='Tab' && !e.shiftKey){ e.preventDefault(); if(hasChips()){ focusFirstChip(); } else { await openMenuSafely(); requestAnimationFrame(()=>{ focusFirstVisibleMenuItem(); }); } } });
                  // Esc inside menu closes and restores label state
                  menu.addEventListener('keydown', (e)=>{ if(e.key==='Escape' && menu.open){ e.preventDefault(); menu.open=false; field.setAttribute('aria-expanded','false'); restoreLabelAfterClose(); field.focus(); } });
                  field.addEventListener('keydown', async (e)=>{
                    if(e.key==='ArrowDown'){ e.preventDefault(); await openMenuSafely(); const first = Array.from(menu.querySelectorAll('md-menu-item')).find(it=>it.style.display!== 'none' && it.hasAttribute('data-code')); if(first) first.focus(); }
                    else if(e.key==='ArrowUp'){ e.preventDefault(); await openMenuSafely(); const items = Array.from(menu.querySelectorAll('md-menu-item')).filter(it=>it.style.display!== 'none' && it.hasAttribute('data-code')); const last=items[items.length-1]; if(last) last.focus(); }
                    else if(e.key==='Enter' || e.key===' '){ e.preventDefault(); await openMenuSafely(); /* keep focus in field to allow typing */ }
                    else if(e.key==='Escape' && menu.open){ e.preventDefault(); menu.open=false; field.setAttribute('aria-expanded','false'); restoreLabelAfterClose(); }
                  });
                  // Use Tab/Shift+Tab to move between visible menu items; close at bounds
                  menu.addEventListener('keydown', (e)=>{
                    if(!menu.open) return;
                    if(e.key==='Tab'){
                      e.preventDefault();
                      const moved = focusRelativeMenuItem(e.shiftKey ? -1 : 1);
                      if(!moved){
                        // At bounds: close and move to next logical control
                        menu.open=false; field.setAttribute('aria-expanded','false'); restoreLabelAfterClose();
                        if(!e.shiftKey){ extractBtn?.focus(); } else { const clearChip = document.getElementById('clearAllChip'); if(clearChip && clearChip.style.display!=='none'){ clearChip.focus(); } else { field.focus(); } }
                      }
                    } else if(e.key===' ' || e.key==='Enter'){
                      const activeItem = e.target && e.target.closest ? e.target.closest('md-menu-item') : null;
                      if(activeItem){
                        e.preventDefault();
                        const code = activeItem.getAttribute('data-code');
                        if(code){
                          const keepTop = menuScrollTop;
                          activeItem.selected = !activeItem.selected;
                          if(activeItem.selected){ selected.add(code); } else { selected.delete(code); }
                          renderChips();
                          requestAnimationFrame(()=>{ menu.scrollTop = keepTop; });
                        } else {
                          // Action item (e.g., Selectează toate)
                          activeItem.click();
                        }
                      }
                    }
                  });
                  field.addEventListener('input', ()=>{ const q=(field.value||''); filterMenu(q); });
                  // From anywhere outside the combo, first Tab focuses the field; second Tab goes to chip (if any) else menu
                  document.addEventListener('keydown', async (e)=>{
                    if(e.key !== 'Tab' || e.shiftKey) return;
                    const active = document.activeElement;
                    const inCombo = active && anchorEl.contains(active);
                    if(!inCombo){ e.preventDefault(); field.focus(); return; }
                    if(active === field){ e.preventDefault(); if(hasChips()){ focusFirstChip(); } else { await openMenuSafely(); requestAnimationFrame(()=>{ focusFirstVisibleMenuItem(); }); } }
                    const clearChip = document.getElementById('clearAllChip');
                    if(active === clearChip){ e.preventDefault(); await openMenuSafely(); requestAnimationFrame(()=>{ focusFirstVisibleMenuItem(); }); }
                  });
                  updateFieldLabel();
                  field.addEventListener('blur', ()=>{ syncLabelFloat(); });
                })();
              </script>
            </body>
            </html>
            """,
            200,
            {"Content-Type": "text/html; charset=utf-8"},
        )

    @app.post("/api/export/jobs")
    def api_export_jobs_start():
        from flask import jsonify
        payload = request.get_json(silent=True) or {}
        codes = payload.get("codes") or []
        labels = payload.get("labels") or []
        list_type = (payload.get("type") or "candidati").strip().lower()
        filename = (payload.get("filename") or "export.xlsx").strip()
        if not codes:
            return {"error": "no codes"}, 400
        job_id = new_job_id()
        jobs[job_id] = {"status":"pending","type":list_type,"done":0,"total":0,"error":"","filename":filename,"result":b""}

        def worker():
            from io import BytesIO
            from openpyxl import Workbook
            from bs4 import BeautifulSoup
            import re as _re
            from .main import fetch_url
            from concurrent.futures import ThreadPoolExecutor, as_completed
            try:
                # Pre-compute totals by discovering page counts for each (code,value)
                # First, collect all values per county
                def get_values(code: str):
                    url = f"https://titularizare.edu.ro/2025/generated/files/j/{code}/index.html"
                    html = fetch_url(url)
                    soup = BeautifulSoup(html, "html.parser")
                    form = soup.find("form", attrs={"name": "paginaJudet"}) or soup
                    sel_id = "repartizare1CandidatiSelect" if list_type == "candidati" else "repartizare1PosturiSelect"
                    sel = form.find("select", id=sel_id) or form.find("select", id=_re.compile(rf"repartizare.*{list_type}", _re.IGNORECASE))
                    vals = []
                    if sel:
                        for opt in sel.find_all("option"):
                            v = (opt.get("value") or "").strip()
                            v = _re.sub(r"/index\.html$", "", v, flags=_re.IGNORECASE).strip()
                            if v:
                                vals.append(v)
                    return vals

                all_pairs: list[tuple[str,str]] = []
                for c in codes:
                    for v in get_values(c):
                        all_pairs.append((c, v))
                # Discover exact last page for each pair concurrently to minimize indeterminate time
                def discover_pair(pair: tuple[str,str]):
                    c, v = pair
                    base = f"https://titularizare.edu.ro/2025/generated/files/j/{c}/{v}"
                    h = fetch_url(f"{base}/index.html")
                    m = _re.findall(r"page_(\d+)\.html", h, flags=_re.IGNORECASE)
                    last = max([int(x) for x in m], default=1)
                    return (pair, last, h)

                last_by_pair: dict[tuple[str,str], int] = {}
                index_html_by_pair: dict[tuple[str,str], str] = {}
                with ThreadPoolExecutor(max_workers=8) as ex:
                    futures = [ex.submit(discover_pair, p) for p in all_pairs]
                    for fut in as_completed(futures):
                        try:
                            pair, last, html0 = fut.result()
                            last_by_pair[pair] = max(last, 1)
                            index_html_by_pair[pair] = html0
                        except Exception:
                            # fallback: assume at least 1 page
                            pass
                total = sum(last_by_pair.get(p, 1) for p in all_pairs)
                jobs[job_id].update({"status":"running","total": total, "done": 0})

                # Build workbook with streaming progress
                wb = Workbook(); first_sheet_used = False
                def safe_title(name: str) -> str:
                    return _re.sub(r"[\\/:\*\?\[\]:]", " ", name).strip()[:31] or "Sheet1"

                def parse_rows(ht):
                    s = BeautifulSoup(ht, "html.parser")
                    t = s.find("table", attrs={"id": "mainTable", "class": _re.compile(r"\bmainTable\b")})
                    if not t:
                        return []
                    out=[]
                    for tr in t.find_all("tr"):
                        cells = [c.get_text(strip=True) for c in tr.find_all(["th","td"]) ]
                        if cells: out.append(cells)
                    return out

                done = 0
                for (c,v) in all_pairs:
                    # county label not strictly needed for title here; keep code
                    title = safe_title(f"{c} - {v}")
                    if not first_sheet_used:
                        ws = wb.active; ws.title = title; first_sheet_used = True
                    else:
                        ws = wb.create_sheet(title=title)
                    base = f"https://titularizare.edu.ro/2025/generated/files/j/{c}/{v}"
                    # index.html first (reuse discovered html when available)
                    try:
                        html0 = index_html_by_pair.get((c,v)) or fetch_url(f"{base}/index.html"); rows = parse_rows(html0)
                        for r in rows: ws.append(r)
                    except Exception:
                        pass
                    done += 1; jobs[job_id]["done"] = done
                    # paginate using discovered last
                    last = max(2, last_by_pair.get((c,v), 1)+1)
                    for p in range(2, last+0):
                        try:
                            hp = fetch_url(f"{base}/page_{p}.html"); pr = parse_rows(hp)
                            if not pr: break
                            for r in pr: ws.append(r)
                            done += 1; jobs[job_id]["done"] = done
                        except Exception:
                            break

                bio = BytesIO(); wb.save(bio); data = bio.getvalue()
                jobs[job_id].update({"status":"ready","result": data})
            except Exception as exc:  # noqa: BLE001
                jobs[job_id].update({"status":"error","error": str(exc)})

        threading.Thread(target=worker, daemon=True).start()
        return {"id": job_id}

    @app.get("/api/export/jobs/<job_id>")
    def api_export_job_status(job_id: str):
        j = jobs.get(job_id)
        if not j:
            return {"error":"not found"}, 404
        return {
            "status": j.get("status"),
            "done": int(j.get("done",0)),
            "total": int(j.get("total",0)),
            "error": j.get("error",""),
            "filename": j.get("filename","export.xlsx"),
        }

    @app.get("/api/export/jobs/<job_id>/download")
    def api_export_job_download(job_id: str):
        j = jobs.get(job_id)
        if not j or j.get("status") != "ready":
            return {"error":"not ready"}, 400
        data: bytes = j.get("result") or b""
        filename = j.get("filename") or "export.xlsx"
        # Content-Disposition safe
        try:
            from urllib.parse import quote as _q
            import re as _re
            ascii_name = filename.encode("ascii", "ignore").decode("ascii") or "export.xlsx"
            ascii_name = _re.sub(r"[\r\n]", " ", ascii_name)
            ascii_name = _re.sub(r"[,;]", " ", ascii_name).strip()
            cd = f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{_q(filename)}"
        except Exception:
            cd = "attachment; filename=export.xlsx"
        headers = {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Disposition": cd,
            "Content-Length": str(len(data)),
        }
        return Response(data, 200, headers)

    @app.get("/api/judete")
    def api_judete():
        from .main import fetch_url
        from bs4 import BeautifulSoup
        import re

        fallback = [
            {"code": "AB", "label": "Alba"}, {"code": "AG", "label": "Argeș"}, {"code": "AR", "label": "Arad"},
            {"code": "B", "label": "București"}, {"code": "BC", "label": "Bacău"}, {"code": "BH", "label": "Bihor"},
            {"code": "BN", "label": "Bistrița-Năsăud"}, {"code": "BR", "label": "Brăila"}, {"code": "BT", "label": "Botoșani"},
            {"code": "BV", "label": "Brașov"}, {"code": "BZ", "label": "Buzău"}, {"code": "CJ", "label": "Cluj"},
            {"code": "CL", "label": "Călărași"}, {"code": "CS", "label": "Caraș-Severin"}, {"code": "CT", "label": "Constanța"},
            {"code": "CV", "label": "Covasna"}, {"code": "DB", "label": "Dâmbovița"}, {"code": "DJ", "label": "Dolj"},
            {"code": "GJ", "label": "Gorj"}, {"code": "GL", "label": "Galați"}, {"code": "GR", "label": "Giurgiu"},
            {"code": "HD", "label": "Hunedoara"}, {"code": "HR", "label": "Harghita"}, {"code": "IF", "label": "Ilfov"},
            {"code": "IL", "label": "Ialomița"}, {"code": "IS", "label": "Iași"}, {"code": "MH", "label": "Mehedinți"},
            {"code": "MM", "label": "Maramureș"}, {"code": "MS", "label": "Mureș"}, {"code": "NT", "label": "Neamț"},
            {"code": "OT", "label": "Olt"}, {"code": "PH", "label": "Prahova"}, {"code": "SB", "label": "Sibiu"},
            {"code": "SJ", "label": "Sălaj"}, {"code": "SM", "label": "Satu Mare"}, {"code": "SV", "label": "Suceava"},
            {"code": "TL", "label": "Tulcea"}, {"code": "TM", "label": "Timiș"}, {"code": "TR", "label": "Teleorman"},
            {"code": "VL", "label": "Vâlcea"}, {"code": "VN", "label": "Vrancea"}, {"code": "VS", "label": "Vaslui"},
        ]

        try:
            html = fetch_url("https://titularizare.edu.ro/2025/generated/files/j/judete.html")
            soup = BeautifulSoup(html, "html.parser")
            areas = soup.select('area[href$="/index.html"]')
            anchors = soup.select('a[href$="/index.html"]')
            items = []
            seen = set()
            pat = re.compile(r"^([A-Z]{1,2})/index\.html$", re.IGNORECASE)

            def normalize_label(raw: str | None) -> str:
                s = (raw or '').strip()
                if not s:
                    return ''
                m = re.search(r"Rapoarte\s+([^\s]+)", s, re.IGNORECASE)
                if m:
                    return m.group(1)
                return s

            def add_item(href, label_text):
                href = (href or '').strip()
                m = pat.match(href)
                if not m:
                    return
                code = m.group(1).upper()
                if code in seen:
                    return
                seen.add(code)
                label = normalize_label(label_text) or code
                items.append({"code": code, "label": label})

            for t in areas:
                add_item(t.get('href'), t.get('alt') or t.get('title'))
            for t in anchors:
                # Prefer text content; fallback to title
                add_item(t.get('href'), t.get_text(strip=True) or t.get('title'))

            if not items:
                items = fallback
            return {"items": items}
        except Exception as exc:
            return {"error": str(exc), "items": fallback}, 200

    @app.get("/api/judete/<code>/candidati-options")
    def api_judet_candidati_options(code: str):
        from .main import fetch_url
        from bs4 import BeautifulSoup
        import re

        # Preserve original case for matching; validate as 1-2 letters
        code = (code or "").strip()
        if not re.fullmatch(r"[A-Za-z]{1,2}", code):
            return {"error": "invalid code", "items": []}, 400

        try:
            # Fetch per-county page; the select lives inside the county page's form
            url = f"https://titularizare.edu.ro/2025/generated/files/j/{code}/index.html"
            html = fetch_url(url)
            soup = BeautifulSoup(html, "html.parser")
            target_form = soup.find("form", attrs={"name": "paginaJudet"}) or soup
            # Inside the form (or whole page as fallback), find the select for candidates
            select = (target_form.find("select", id="repartizare1CandidatiSelect") or
                      target_form.find("select", id=re.compile(r"repartizare.*Candidati", re.IGNORECASE)))
            if not select:
                return {"items": []}
            items = []
            for opt in select.find_all("option"):
                value = (opt.get("value") or "").strip()
                # Normalize by removing trailing /index.html from relative paths
                value = re.sub(r"/index\.html$", "", value, flags=re.IGNORECASE).strip()
                label = (opt.get_text(strip=True) or "").strip()
                if not value:
                    continue
                items.append({"value": value, "label": label})
            return {"items": items}
        except Exception as exc:  # noqa: BLE001
            return {"error": str(exc), "items": []}, 200

    @app.get("/api/judete/<code>/posturi-options")
    def api_judet_posturi_options(code: str):
        from .main import fetch_url
        from bs4 import BeautifulSoup
        import re

        code = (code or "").strip()
        if not re.fullmatch(r"[A-Za-z]{1,2}", code):
            return {"error": "invalid code", "items": []}, 400

        try:
            url = f"https://titularizare.edu.ro/2025/generated/files/j/{code}/index.html"
            html = fetch_url(url)
            soup = BeautifulSoup(html, "html.parser")
            target_form = soup.find("form", attrs={"name": "paginaJudet"}) or soup
            select = (target_form.find("select", id="repartizare1PosturiSelect") or
                      target_form.find("select", id=re.compile(r"repartizare.*Posturi", re.IGNORECASE)))
            if not select:
                return {"items": []}
            items = []
            for opt in select.find_all("option"):
                value = (opt.get("value") or "").strip()
                # Normalize by removing trailing /index.html from relative paths
                value = re.sub(r"/index\.html$", "", value, flags=re.IGNORECASE).strip()
                label = (opt.get_text(strip=True) or "").strip()
                if not value:
                    continue
                items.append({"value": value, "label": label})
            return {"items": items}
        except Exception as exc:  # noqa: BLE001
            return {"error": str(exc), "items": []}, 200

    @app.get("/api/export/xlsx")
    def api_export_xlsx():
        from io import BytesIO
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        # Optional sheet title and filename from query params
        raw_title = (request.args.get("sheet") or "").strip()
        def sanitize_sheet_name(name: str) -> str:
            import re
            name = re.sub(r"[\\/:\*\?\[\]:]", " ", name)
            name = name.strip()[:31]
            return name or "Sheet1"
        ws.title = sanitize_sheet_name(raw_title)
        # Leave empty per requirements
        bio = BytesIO()
        wb.save(bio)
        data = bio.getvalue()
        download_name = (request.args.get("filename") or "export.xlsx").strip() or "export.xlsx"
        # Safe Content-Disposition: avoid commas and special chars in ASCII filename
        try:
            from urllib.parse import quote as _q
            import re as _re
            ascii_name = download_name.encode("ascii", "ignore").decode("ascii") or "export.xlsx"
            ascii_name = _re.sub(r"[\r\n]", " ", ascii_name)
            ascii_name = _re.sub(r"[,;]", " ", ascii_name).strip()
            cd = f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{_q(download_name)}"
        except Exception:
            cd = "attachment; filename=export.xlsx"
        headers = {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Disposition": cd,
            "Content-Length": str(len(data)),
        }
        return Response(data, 200, headers)

    @app.get("/api/export/judet")
    def api_export_judet():
        from io import BytesIO
        from openpyxl import Workbook
        from bs4 import BeautifulSoup
        import re
        # Reuse shared fetch with headers/timeouts
        from .main import fetch_url

        code = (request.args.get("code") or "").strip().upper()
        value = (request.args.get("value") or "").strip()
        sheet = (request.args.get("sheet") or "Sheet1").strip()
        filename = (request.args.get("filename") or "export.xlsx").strip()

        if not re.fullmatch(r"[A-Z]{1,2}", code):
            return {"error": "invalid code"}, 400
        # Ensure value has no trailing /index.html
        value = re.sub(r"/index\.html$", "", value, flags=re.IGNORECASE).strip()
        if not value:
            return {"error": "invalid value"}, 400

        base = f"https://titularizare.edu.ro/2025/generated/files/j/{code}/{value}"
        # Fetch first page
        try:
            html = fetch_url(f"{base}/index.html")
        except Exception as exc:
            return {"error": str(exc)}, 400

        def parse_rows(html_text: str):
            s = BeautifulSoup(html_text, "html.parser")
            t = s.find("table", attrs={"id": "mainTable", "class": re.compile(r"\bmainTable\b")})
            out = []
            if not t:
                return out
            for tr in t.find_all("tr"):
                cells = [c.get_text(strip=True) for c in tr.find_all(["th", "td"]) ]
                if cells:
                    out.append(cells)
            return out

        rows = parse_rows(html)

        # Discover pagination links like page_2.html, page_3.html and compute max page
        soup = BeautifulSoup(html, "html.parser")
        page_nums = set()
        # Collect from anchors
        for a in soup.find_all("a", href=True):
            m = re.search(r"page_(\d+)\.html", a["href"], re.IGNORECASE)
            if m:
                try:
                    page_nums.add(int(m.group(1)))
                except Exception:
                    pass
        # Also collect from raw HTML (scripts like drawLastButton('page_33.html'))
        for n in re.findall(r"page_(\d+)\.html", html, flags=re.IGNORECASE):
            try:
                page_nums.add(int(n))
            except Exception:
                pass

        # Robust pagination: attempt sequential pages until a fetch fails (404) or table missing
        p = 2
        while True:
            try:
                html_p = fetch_url(f"{base}/page_{p}.html")
                parsed = parse_rows(html_p)
                if not parsed:
                    break
                rows.extend(parsed)
                p += 1
            except Exception:
                break

        # Build workbook
        wb = Workbook()
        ws = wb.active
        # Sanitize sheet title
        safe_title = re.sub(r"[\\/:\*\?\[\]:]", " ", sheet).strip()[:31] or "Sheet1"
        ws.title = safe_title
        for r in rows:
            ws.append(r)

        bio = BytesIO()
        wb.save(bio)
        data = bio.getvalue()
        # Robust Content-Disposition with ASCII fallback + RFC 5987 filename*
        try:
            from urllib.parse import quote as _q
            ascii_name = filename.encode("ascii", "ignore").decode("ascii") or "export.xlsx"
            cd = f"attachment; filename={ascii_name}; filename*=UTF-8''{_q(filename)}"
        except Exception:
            cd = f"attachment; filename=export.xlsx"
        headers = {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Disposition": cd,
            "Content-Length": str(len(data)),
        }
        return Response(data, 200, headers)

    @app.get("/api/export/judet-bulk")
    def api_export_judet_bulk():
        # Build a single workbook with multiple sheets, one per list item value for the given county
        from io import BytesIO
        from openpyxl import Workbook
        from bs4 import BeautifulSoup
        import re
        from .main import fetch_url

        code = (request.args.get("code") or "").strip().upper()
        list_type = (request.args.get("type") or "candidati").strip().lower()
        sheet_base = (request.args.get("sheetBase") or code).strip()
        filename = (request.args.get("filename") or f"export_all_{code}.xlsx").strip()

        if not re.fullmatch(r"[A-Z]{1,2}", code):
            return {"error": "invalid code"}, 400

        # Determine options endpoint
        opt_endpoint = f"/api/judete/{code}/candidati-options" if list_type == "candidati" else f"/api/judete/{code}/posturi-options"
        try:
            # Call our own API to get normalized values for this county
            # Use internal fetch_url to avoid CORS; build absolute URL
            # Reuse server base since we are in the same app
            from urllib.parse import urljoin
            # Assuming local host/port; use relative path via requests within same process is not trivial without server URL
            # Fallback: reconstruct data by fetching the county page directly (same logic as endpoints)
            county_url = f"https://titularizare.edu.ro/2025/generated/files/j/{code}/index.html"
            html = fetch_url(county_url)
            soup = BeautifulSoup(html, "html.parser")
            form = soup.find("form", attrs={"name": "paginaJudet"}) or soup
            select_id = "repartizare1CandidatiSelect" if list_type == "candidati" else "repartizare1PosturiSelect"
            select = form.find("select", id=select_id) or form.find("select", id=re.compile(rf"repartizare.*{list_type}", re.IGNORECASE))
            values = []
            if select:
                for opt in select.find_all("option"):
                    v = (opt.get("value") or "").strip()
                    v = re.sub(r"/index\.html$", "", v, flags=re.IGNORECASE).strip()
                    if v:
                        values.append({"value": v, "label": (opt.get_text(strip=True) or "").strip()})
        except Exception as exc:
            return {"error": str(exc)}, 400

        # Build workbook with multiple sheets
        wb = Workbook()
        # First sheet will be overwritten; track if we used it
        first_sheet_used = False

        def add_sheet_for_value(val: str, label: str):
            base = f"https://titularizare.edu.ro/2025/generated/files/j/{code}/{val}"
            try:
                html0 = fetch_url(f"{base}/index.html")
            except Exception:
                return
            rows_all = []
            def parse_rows(ht):
                s = BeautifulSoup(ht, "html.parser")
                t = s.find("table", attrs={"id": "mainTable", "class": re.compile(r"\bmainTable\b")})
                if not t:
                    return []
                out=[]
                for tr in t.find_all("tr"):
                    cells = [c.get_text(strip=True) for c in tr.find_all(["th","td"]) ]
                    if cells:
                        out.append(cells)
                return out
            rows_all.extend(parse_rows(html0))
            # paginate sequentially
            p=2
            while True:
                try:
                    h = fetch_url(f"{base}/page_{p}.html")
                    pr = parse_rows(h)
                    if not pr:
                        break
                    rows_all.extend(pr)
                    p += 1
                except Exception:
                    break
            # Create sheet and append rows
            import re as _re
            title = f"{sheet_base} - {val}"
            title = _re.sub(r"[\\/:\*\?\[\]:]", " ", title).strip()[:31] or "Sheet1"
            nonlocal first_sheet_used
            if not first_sheet_used:
                ws = wb.active
                ws.title = title
                first_sheet_used = True
            else:
                ws = wb.create_sheet(title=title)
            for r in rows_all:
                ws.append(r)

        for it in values:
            add_sheet_for_value(it["value"], it.get("label", ""))

        bio = BytesIO()
        wb.save(bio)
        data = bio.getvalue()
        # Robust Content-Disposition with ASCII fallback + RFC 5987 filename*
        try:
            from urllib.parse import quote as _q
            ascii_name = filename.encode("ascii", "ignore").decode("ascii") or "export.xlsx"
            cd = f"attachment; filename={ascii_name}; filename*=UTF-8''{_q(filename)}"
        except Exception:
            cd = f"attachment; filename=export.xlsx"
        headers = {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Disposition": cd,
            "Content-Length": str(len(data)),
        }
        return Response(data, 200, headers)

    @app.get("/api/export/judete-multi-bulk")
    def api_export_judete_multi_bulk():
        # Export multiple selected counties into a single workbook; each sheet corresponds to one list item per county
        from io import BytesIO
        from openpyxl import Workbook
        from bs4 import BeautifulSoup
        import re
        from .main import fetch_url

        # Inputs
        codes_csv = (request.args.get("codes") or "").strip()
        list_type = (request.args.get("type") or "candidati").strip().lower()
        labels_joined = (request.args.get("labels") or "").strip()
        filename = (request.args.get("filename") or "export_mixt.xlsx").strip()

        # Parse inputs
        codes = [c.strip().upper() for c in codes_csv.split(',') if c.strip()]
        if not codes:
            return {"error": "no codes"}, 400
        county_labels = labels_joined.split('|') if labels_joined else codes

        # Helper to get options for a county
        def get_values_for_county(code: str):
            url = f"https://titularizare.edu.ro/2025/generated/files/j/{code}/index.html"
            html = fetch_url(url)
            soup = BeautifulSoup(html, "html.parser")
            form = soup.find("form", attrs={"name": "paginaJudet"}) or soup
            sel_id = "repartizare1CandidatiSelect" if list_type == "candidati" else "repartizare1PosturiSelect"
            sel = form.find("select", id=sel_id) or form.find("select", id=re.compile(rf"repartizare.*{list_type}", re.IGNORECASE))
            out = []
            if sel:
                for opt in sel.find_all("option"):
                    v = (opt.get("value") or "").strip()
                    v = re.sub(r"/index\.html$", "", v, flags=re.IGNORECASE).strip()
                    if v:
                        out.append({"value": v, "label": (opt.get_text(strip=True) or "").strip()})
            return out

        # Helper to fetch all pages for a base value
        def fetch_rows(code: str, val: str):
            base = f"https://titularizare.edu.ro/2025/generated/files/j/{code}/{val}"
            def parse_rows(ht):
                s = BeautifulSoup(ht, "html.parser")
                t = s.find("table", attrs={"id": "mainTable", "class": re.compile(r"\bmainTable\b")})
                if not t:
                    return []
                out = []
                for tr in t.find_all("tr"):
                    cells = [c.get_text(strip=True) for c in tr.find_all(["th","td"]) ]
                    if cells:
                        out.append(cells)
                return out
            rows_all = []
            try:
                html0 = fetch_url(f"{base}/index.html")
                rows_all.extend(parse_rows(html0))
                p = 2
                while True:
                    try:
                        hp = fetch_url(f"{base}/page_{p}.html")
                        pr = parse_rows(hp)
                        if not pr:
                            break
                        rows_all.extend(pr)
                        p += 1
                    except Exception:
                        break
            except Exception:
                pass
            return rows_all

        # Build workbook
        wb = Workbook()
        first_sheet_used = False

        import re as _re
        def safe_title(name: str) -> str:
            return _re.sub(r"[\\/:\*\?\[\]:]", " ", name).strip()[:31] or "Sheet1"

        # Iterate counties and their values
        for idx, code in enumerate(codes):
            county_label = county_labels[idx] if idx < len(county_labels) else code
            values = get_values_for_county(code)
            for it in values:
                val = it["value"]
                rows = fetch_rows(code, val)
                title = safe_title(f"{county_label} - {val}")
                if not first_sheet_used:
                    ws = wb.active
                    ws.title = title
                    first_sheet_used = True
                else:
                    ws = wb.create_sheet(title=title)
                for r in rows:
                    ws.append(r)

        bio = BytesIO()
        wb.save(bio)
        data = bio.getvalue()
        # Robust Content-Disposition with ASCII fallback; avoid commas/CRLF
        try:
            from urllib.parse import quote as _q
            import re as _re
            ascii_name = filename.encode("ascii", "ignore").decode("ascii") or "export.xlsx"
            ascii_name = _re.sub(r"[\r\n]", " ", ascii_name)
            ascii_name = _re.sub(r"[,;]", " ", ascii_name).strip()
            cd = f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{_q(filename)}"
        except Exception:
            cd = "attachment; filename=export.xlsx"
        headers = {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Disposition": cd,
            "Content-Length": str(len(data)),
        }
        return Response(data, 200, headers)

    return app


__all__ = ["create_app"]


